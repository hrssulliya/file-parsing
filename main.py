from flask import Flask, jsonify, request
from flask_sqlalchemy import SQLAlchemy
from openpyxl import  load_workbook, Workbook
from openpyxl.styles import PatternFill, GradientFill, fonts
from openpyxl.formatting.rule import CellIsRule
import time


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///store.db"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)



class Product(db.Model):
    variant_id = db.Column(db.Integer, primary_key=True)
    category_name = db.Column(db.String(200), nullable=False)
    variant_name = db.Column(db.String(200), nullable=False)
    current_price = db.Column(db.Integer, nullable=False)


@app.route('/',methods=["POST"])
def test():
    req_data = request.files['file']
    # startScript = time.time()
    workbook = load_workbook(filename=req_data)
    sheet = workbook.active
    for value in sheet.iter_rows(min_row=2,values_only=True):
        variant_id = value[0]
        category_name = value[1]
        variant_name = value[2]
        current_price = value[3]
        # store = Product(variant_id=variant_id, category_name=category_name, variant_name=variant_name, current_price=current_price)
        # db.session.add(store)
        if Product.query.filter_by(variant_id=value[0]).first():
            continue
        else:
            store = Product(variant_id=variant_id, category_name=category_name, variant_name=variant_name, current_price=current_price)
            db.session.add(store)
        
        db.session.commit()
    return jsonify ({'massage': 'excel file uplode'})
    # endScript = time.time()
    # time_of_Script = endScript - startScript
    # print(str(time_of_Script) + 'second')



# @app.route('/edit',methods=["PUT"])
# def edit():
#     req_data = request.files['file']
#     # startScript = time.time()
#     workbook = load_workbook(filename=req_data)
#     sheet = workbook.active
#     for value in sheet.iter_rows(min_row=2,values_only=True):
#         variant_id = value[0]
#         category_name = value[1]
#         variant_name = value[2]
#         current_price = value[3]
#         store = Product.query.filter_by(variant_id=value[0]).first()
#         if len(store) <1:
#             store.category_name = category_name 
#             store.variant_name = variant_name 
#             store.current_price = current_price
#             db.session.add(store)
#         else:
#             store = Product(variant_id=variant_id, category_name=category_name, variant_name=variant_name, current_price=current_price)
#             db.session.add(store)
#         db.session.commit()
#     return jsonify ({'massage': 'excel file update'})
#     # endScript = time.time()
#     # time_of_Script = endScript - startScript
#     # print(str(time_of_Script) + 'second')


@app.route('/download/<string:name>')
def downlode(name):
    fileName = name+'.xlsx'
    book = Workbook()

    sheet = book.active
    sheet.title = 'Data'
    sheet.append(['variant_id','category_name','variant_name','current_price'])
    book.save(fileName)
    allProduct = Product.query.all()

    for product in allProduct:
        item = product.variant_id,product.category_name,product.variant_name,product.current_price
        row = list(item)
        book = load_workbook(fileName)
        sheet = book.active
        sheet.append(row)
        
        book.save(fileName)
    book = load_workbook(fileName)
    sheet = book.active
    redfill = PatternFill(start_color='EE4636',end_color='EE4636',fill_type='solid')
    sheet.conditional_formatting.add('D2:D400',CellIsRule(operator='greaterThan',formula=['100'],stopIfTrue=True,fill=redfill))

    book.save(fileName)
    return jsonify ({'massage': f'{fileName} file download'})

with app.app_context():
    db.create_all()


app.run(debug=True, port=5600)

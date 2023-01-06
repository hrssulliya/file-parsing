from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, GradientFill

workbook = load_workbook(filename="Menu2.xlsx")
sheet = workbook.active

# for value in sheet.iter_rows(min_row=1,values_only=True):
#     print(value)

column_a = sheet['D']
for cell in column_a:
    if cell >= 100:
        cell.fill = PatternFill('solid', fgColor = 'F2F2F2')


# rows = sheet['2']
# for row in rows:
#     if row == sheet['A1']:
#         continue
#     else:
#         print(row,end=',')


# column_a = sheet['A']
# count = 1
# for cell in column_a:
#     rows = sheet[count]
#     for row in rows:
#         if cell == sheet['A1']:
#             continue
#         else:
#             print(cell+str(row).value,end=',')
#         count+=1

# book = Workbook()
# sheet = book.active
# sheet.title = 'Data'
# sheet.append(['variant_id','category_name','variant_name','current_price'])
# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )

# for row in rows:
#     sheet.append(row)

# book.save('appending.xlsx')

# fileName = 'appending.xlsx'
# book = load_workbook(fileName)
# sheet = book.active
# sheet.title = 'Data'
# # sheet.append([418575787, 'ice-creem', 'lassi', 35])
# rows = (
#     (88, 46, 57),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )

# for row in rows:
#     sheet.append(row)

# book.save(fileName)